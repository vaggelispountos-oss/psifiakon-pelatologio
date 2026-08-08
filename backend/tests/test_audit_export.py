"""
test_audit_export.py
--------------------------------------------------------------------
Τεστ για το /api/audit-log/export.xlsx (εξαγωγή για λογιστικό έλεγχο).

Δεν αρκεί να ελεγχθεί ότι γυρνά 200: το ζητούμενο είναι ότι το αρχείο
ΑΝΟΙΓΕΙ όντως ως Excel και περιέχει τα σωστά δεδομένα ΜΟΝΟ του συνεργείου
που το ζήτησε. Γι' αυτό κάθε test ξαναδιαβάζει το παραγόμενο αρχείο με
openpyxl, όπως θα έκανε το Excel του λογιστή.

ΚΑΘΑΡΙΖΕΙ μετά τον εαυτό του ώστε να μη μολύνει τη dev βάση.
--------------------------------------------------------------------
"""
import io
import os
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import app  # noqa: E402
from models import AadeLog, DclEntry, Workshop, db  # noqa: E402

PASSWORD = "testpass1234"
EMAIL_A = "auditexport.a@example.com"
EMAIL_B = "auditexport.b@example.com"


def client():
    app.config["TESTING"] = True
    return app.test_client()


def _register(c, email):
    res = c.post(
        "/api/auth/register",
        json={
            "name": f"Test {email}",
            "email": email,
            "password": PASSWORD,
            "businessType": "garage",
            "termsAccepted": True,
        },
    )
    if res.status_code not in (200, 201):
        res = c.post("/api/auth/login", json={"email": email, "password": PASSWORD})
    data = res.get_json()
    return data["accessToken"], data["workshop"]["id"]


@pytest.fixture
def two_workshops():
    """
    Δύο ξεχωριστά συνεργεία, το καθένα με μία εγγραφή + ένα AadeLog.
    Επιστρέφει (token_a, token_b) και καθαρίζει ΟΛΑ όσα δημιούργησε.
    """
    c = client()
    with app.app_context():
        token_a, ws_a = _register(c, EMAIL_A)
        token_b, ws_b = _register(c, EMAIL_B)

        created = []
        for ws_id, plate in ((ws_a, "AAA1111"), (ws_b, "BBB2222")):
            entry = DclEntry(
                workshop_id=ws_id, plate=plate, branch=0, status="open", id_dcl=f"DCL-{plate}"
            )
            db.session.add(entry)
            db.session.flush()
            log = AadeLog(
                workshop_id=ws_id,
                dcl_entry_id=entry.id,
                method="SendClient",
                request_json='{"plate": "%s"}' % plate,
                response_json='{"idDcl": "DCL-%s"}' % plate,
                success=True,
            )
            db.session.add(log)
            created.append((entry, log))
        db.session.commit()

        yield token_a, token_b

        for entry, log in created:
            db.session.delete(log)
            db.session.delete(entry)
        for email in (EMAIL_A, EMAIL_B):
            ws = Workshop.query.filter_by(email=email).first()
            if ws:
                AadeLog.query.filter_by(workshop_id=ws.id).delete()
                DclEntry.query.filter_by(workshop_id=ws.id).delete()
                db.session.delete(ws)
        db.session.commit()


def _download(token):
    res = client().get(
        "/api/audit-log/export.xlsx", headers={"Authorization": f"Bearer {token}"}
    )
    assert res.status_code == 200
    return res


def test_requires_authentication():
    res = client().get("/api/audit-log/export.xlsx")
    assert res.status_code == 401


def test_returns_a_real_xlsx_file(two_workshops):
    token_a, _ = two_workshops
    res = _download(token_a)

    assert "spreadsheetml.sheet" in res.headers["Content-Type"]
    assert "attachment" in res.headers["Content-Disposition"]
    assert ".xlsx" in res.headers["Content-Disposition"]

    # Το ουσιαστικό: ανοίγει όντως ως Excel;
    wb = load_workbook(io.BytesIO(res.data))
    assert wb.sheetnames == ["Εγγραφές", "Κλήσεις ΑΑΔΕ"]


def test_contains_own_data(two_workshops):
    token_a, _ = two_workshops
    wb = load_workbook(io.BytesIO(_download(token_a).data))

    entries_text = "\n".join(
        str(c.value) for row in wb["Εγγραφές"].iter_rows() for c in row
    )
    assert "AAA1111" in entries_text
    assert "DCL-AAA1111" in entries_text

    logs_text = "\n".join(
        str(c.value) for row in wb["Κλήσεις ΑΑΔΕ"].iter_rows() for c in row
    )
    assert "SendClient" in logs_text


def test_does_not_leak_other_workshops_data(two_workshops):
    """
    Multi-tenant: το συνεργείο Α δεν πρέπει να δει ΤΙΠΟΤΑ του Β — ούτε
    πινακίδα, ούτε idDcl, ούτε περιεχόμενο κλήσης ΑΑΔΕ.
    """
    token_a, _ = two_workshops
    everything = "\n".join(
        str(c.value)
        for sheet in load_workbook(io.BytesIO(_download(token_a).data)).worksheets
        for row in sheet.iter_rows()
        for c in row
    )
    assert "BBB2222" not in everything
    assert "DCL-BBB2222" not in everything


def test_oversized_and_illegal_content_does_not_corrupt_the_file():
    """
    Τα request/response της ΑΑΔΕ ξεπερνούν άνετα το όριο των 32.767
    χαρακτήρων ανά κελί, και μπορεί να περιέχουν control chars που το XML
    του .xlsx δεν δέχεται. Και τα δύο παράγουν αρχείο που το Excel ΑΡΝΕΙΤΑΙ
    να ανοίξει — δηλαδή αθόρυβη αποτυχία ακριβώς τη μέρα του ελέγχου.
    Δοκιμάζει τον builder απευθείας, χωρίς βάση.
    """
    from datetime import datetime

    from audit_export import build_audit_workbook

    class FakeEntry:
        id = 1
        plate = "ZZZ9999"
        id_dcl = "DCL-1"
        status = "open"
        creation_date_time = datetime(2026, 8, 8, 14, 30)
        completion_date_time = None
        provided_service_category = 5
        provided_service_category_other = "Ειδική εργασία"
        invoice_kind = 1
        reason_non_issue_type = None
        mark = "400001"
        correlate_id = "C1"
        amount = 123.45
        vehicle_movement_purpose = None
        created_by_employee = None

    class FakeLog:
        created_at = datetime(2026, 8, 8, 14, 31)
        dcl_entry_id = 1
        dcl_entry = FakeEntry()
        method = "SendClient"
        success = True
        actor_employee = None
        request_json = "X" * 100000 + "\x00\x01\x0b"
        response_json = "Y" * 50000

    stream = build_audit_workbook([FakeEntry()], [FakeLog()])
    wb = load_workbook(io.BytesIO(stream.getvalue()))

    cell = wb["Κλήσεις ΑΑΔΕ"]["G2"].value
    assert len(cell) <= 32767, "ξεπερνά το όριο χαρακτήρων κελιού του Excel"
    assert cell.endswith("[κόπηκε]")
    assert "\x00" not in cell and "\x0b" not in cell

    # Η κατηγορία «Λοιπά» πρέπει να συνοδεύεται από το ελεύθερο κείμενο,
    # αλλιώς ο ελεγκτής βλέπει «Λοιπά» και δεν μαθαίνει τίποτα.
    assert wb["Εγγραφές"]["F2"].value == "Λοιπά: Ειδική εργασία"
    # Ημερομηνίες ως πραγματικά date cells (ταξινόμηση/φιλτράρισμα στο Excel)
    assert isinstance(wb["Εγγραφές"]["E2"].value, datetime)


def test_headers_are_present_even_with_no_data(two_workshops):
    """
    Νέος λογαριασμός χωρίς εγγραφές: το αρχείο πρέπει να έχει κεφαλίδες και
    να ανοίγει κανονικά, όχι να είναι κενό/κατεστραμμένο.
    """
    c = client()
    token, _ = _register(c, "auditexport.empty@example.com")
    try:
        wb = load_workbook(io.BytesIO(_download(token).data))
        assert wb["Εγγραφές"]["A1"].value == "ID"
        assert wb["Κλήσεις ΑΑΔΕ"]["A1"].value == "Ημερομηνία"
    finally:
        with app.app_context():
            ws = Workshop.query.filter_by(email="auditexport.empty@example.com").first()
            if ws:
                db.session.delete(ws)
                db.session.commit()
