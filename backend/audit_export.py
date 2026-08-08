"""
audit_export.py
--------------------------------------------------------------------
Παραγωγή αρχείου Excel (.xlsx) με το ιστορικό εγγραφών + το audit trail των
κλήσεων προς την ΑΑΔΕ, για λογιστικό έλεγχο / φακέλωση.

Διαφορά από το GDPR export (`/api/account/export`, routes_account.py): εκείνο
βγάζει ΟΛΑ τα δεδομένα του λογαριασμού σε JSON για φορητότητα· αυτό εδώ βγάζει
ΜΟΝΟ ό,τι χρειάζεται ένας λογιστής/ελεγκτής, σε μορφή που ανοίγει σε Excel.

Οι ελληνικές ετικέτες των κωδικών ΑΑΔΕ ζούσαν μέχρι τώρα ΜΟΝΟ στο frontend
(constants.js). Επαναλαμβάνονται εδώ επειδή το αρχείο παράγεται server-side —
αν αλλάξει κωδικός ΑΑΔΕ, πρέπει να ενημερωθούν ΚΑΙ ΤΑ ΔΥΟ σημεία.
--------------------------------------------------------------------
"""
import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Ένα κελί Excel δεν χωρά πάνω από 32.767 χαρακτήρες — τα request/response
# JSON της ΑΑΔΕ ξεπερνούν άνετα αυτό το όριο. Χωρίς κόψιμο, το openpyxl
# παράγει αρχείο που το Excel αρνείται να ανοίξει.
MAX_CELL_CHARS = 32000
_TRUNCATED = "… [κόπηκε]"

# Χαρακτήρες ελέγχου που το XML του .xlsx δεν δέχεται. Μπορούν να βρεθούν σε
# raw responses της ΑΑΔΕ και παράγουν κατεστραμμένο αρχείο.
_ILLEGAL_CHARS = re.compile(r"[\000-\010\013\014\016-\037]")

SERVICE_CATEGORIES = {
    1: "Εργασία με χρήση ανταλλακτικών",
    2: "Εργασία με ανταλλακτικά που φέρνει ο πελάτης",
    3: "Εργασία χωρίς ανταλλακτικά",
    4: "Δωρεάν υπηρεσία",
    5: "Λοιπά",
    6: "Αποζημίωση παροχής εγγύησης",
    9: "Ιδιόχρηση",
}

INVOICE_KINDS = {
    1: "ΑΛΠ / ΑΠΥ",
    2: "Τιμολόγιο",
    3: "ΑΛΠ / ΑΠΥ - ΦΗΜ",
}

REASON_NON_ISSUE_TYPES = {
    1: "Δωρεάν Υπηρεσία",
    2: "Ιδιόχρηση",
    3: "Αποζημίωση Παροχής Εγγύησης",
}

VEHICLE_MOVEMENT_PURPOSES = {
    1: "Ενοικίαση",
    2: "Ιδιόχρηση",
    3: "Δωρεάν Υπηρεσία",
}

STATUS_LABELS = {
    "open": "Μπήκε / Έξω",
    "in_progress": "Σε εξέλιξη",
    "completed": "Ολοκληρωμένη",
    "correlated": "Συσχετισμένη",
    "cancelled": "Ακυρωμένη",
}

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_DATE_FORMAT = "DD/MM/YYYY HH:MM"


def _clean(value):
    """str -> ασφαλές για κελί Excel (κομμένο + χωρίς control chars)."""
    if value is None:
        return None
    text = str(value)
    text = _ILLEGAL_CHARS.sub("", text)
    if len(text) > MAX_CELL_CHARS:
        text = text[: MAX_CELL_CHARS - len(_TRUNCATED)] + _TRUNCATED
    return text


def _label(mapping, value):
    if value is None:
        return None
    return mapping.get(int(value), f"Άγνωστος κωδικός ({value})")


def _actor_name(employee):
    # None = ο ιδιοκτήτης (δεν έχει Employee row) ή εγγραφή από πριν μπει η
    # στήλη — δες models.AadeLog.actor_employee_id.
    return employee.name if employee else "Ιδιοκτήτης"


def _header_row(ws, titles):
    cells = []
    for title in titles:
        cell = WriteOnlyCell(ws, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        cells.append(cell)
    ws.append(cells)


def _row(ws, values):
    """
    Γράφει γραμμή μετατρέποντας τα datetime σε πραγματικά κελιά ημερομηνίας
    (με ελληνικό format) αντί για κείμενο — ώστε να ταξινομούνται και να
    φιλτράρονται σωστά μέσα στο Excel.
    """
    cells = []
    for value in values:
        if isinstance(value, datetime):
            cell = WriteOnlyCell(ws, value=value)
            cell.number_format = _DATE_FORMAT
        else:
            cell = WriteOnlyCell(ws, value=_clean(value) if isinstance(value, str) else value)
        cells.append(cell)
    ws.append(cells)


def _set_widths(ws, widths):
    # Πρέπει να οριστεί ΠΡΙΝ γραφτούν γραμμές σε write_only mode.
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def build_audit_workbook(entries, logs):
    """
    entries: λίστα DclEntry (με φορτωμένο created_by_employee)
    logs:    λίστα AadeLog (με φορτωμένα actor_employee και dcl_entry)
    Επιστρέφει BytesIO έτοιμο για send_file.

    write_only=True: το openpyxl κρατά τις γραμμές στη μνήμη σε κανονικό mode.
    Ένα συνεργείο με μερικές χιλιάδες εγγραφές × ολόκληρα XML responses θα
    έφτανε εύκολα το όριο μνήμης του Render free plan (512MB).
    """
    wb = Workbook(write_only=True)

    # --- Φύλλο 1: Εγγραφές ------------------------------------------------
    ws = wb.create_sheet("Εγγραφές")
    _set_widths(ws, [8, 12, 18, 16, 18, 30, 18, 18, 22, 18, 14, 12, 16, 30])
    ws.freeze_panes = "A2"
    _header_row(
        ws,
        [
            "ID",
            "Πινακίδα",
            "idDcl (ΑΑΔΕ)",
            "Κατάσταση",
            "1ος Χρόνος",
            "Κατηγορία υπηρεσίας",
            "3ος Χρόνος",
            "Είδος παραστατικού",
            "Αιτιολογία μη έκδοσης",
            "ΜΑΡΚ",
            "correlateId",
            "Ποσό (€)",
            "Σκοπός κίνησης",
            "Υπάλληλος",
        ],
    )
    for e in entries:
        category = _label(SERVICE_CATEGORIES, e.provided_service_category)
        # Η κατηγορία «Λοιπά» έχει υποχρεωτικό ελεύθερο κείμενο — χωρίς αυτό
        # ο ελεγκτής βλέπει σκέτο «Λοιπά» και δεν μαθαίνει τίποτα.
        if e.provided_service_category_other:
            category = f"{category or 'Λοιπά'}: {e.provided_service_category_other}"
        _row(
            ws,
            [
                e.id,
                e.plate,
                e.id_dcl,
                STATUS_LABELS.get(e.status, e.status),
                e.creation_date_time,
                category,
                e.completion_date_time,
                _label(INVOICE_KINDS, e.invoice_kind),
                _label(REASON_NON_ISSUE_TYPES, e.reason_non_issue_type),
                e.mark,
                e.correlate_id,
                e.amount,
                _label(VEHICLE_MOVEMENT_PURPOSES, e.vehicle_movement_purpose),
                _actor_name(e.created_by_employee),
            ],
        )

    # --- Φύλλο 2: Audit trail ΑΑΔΕ ---------------------------------------
    ws2 = wb.create_sheet("Κλήσεις ΑΑΔΕ")
    _set_widths(ws2, [18, 10, 12, 22, 12, 18, 60, 60])
    ws2.freeze_panes = "A2"
    _header_row(
        ws2,
        [
            "Ημερομηνία",
            "ID εγγραφής",
            "Πινακίδα",
            "Μέθοδος ΑΑΔΕ",
            "Επιτυχία",
            "Υπάλληλος",
            "Αίτημα (JSON)",
            "Απάντηση (JSON)",
        ],
    )
    for log in logs:
        _row(
            ws2,
            [
                log.created_at,
                log.dcl_entry_id,
                log.dcl_entry.plate if log.dcl_entry else None,
                log.method,
                "ΝΑΙ" if log.success else "ΟΧΙ",
                _actor_name(log.actor_employee),
                log.request_json,
                log.response_json,
            ],
        )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
